#!/usr/bin/env python3
"""
Saha Durum ve Operasyon Detay Tablosu (xlsx) -> dashboard-data.json

Kullanim:
    python parse_saha_durum.py <xlsx_yolu> <json_cikti_yolu>

Varsayilanlar:
    xlsx_yolu       = saha_durum_operasyon_tablosu.xlsx
    json_cikti_yolu = data/dashboard-data.json

Bu script, ACEP Saha Operasyon Dashboard HTML sayfasindaki client-side
JS parser (parseWorkbook) ile birebir ayni mantigi kullanir; boylece
ister tarayicida yerel xlsx yuklensin, ister GitHub Actions bu scripti
calistirsin, ayni veri yapisi uretilir.
"""

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("HATA: openpyxl kurulu degil. `pip install openpyxl` calistirin.", file=sys.stderr)
    sys.exit(1)

STAGE_ORDER = [
    "Keşif Bekleyen",
    "Altyapı Planlanan",
    "Altyapı Devam Eden",
    "Altyapı Tamamlanan",
    "Sevkiyat Yapılan",
    "Kurulumu Devam Eden",
    "Kurulumu Tamamlanan",
    "Devreye Alınan",
]


_TR_LOWER_MAP = str.maketrans({"İ": "i", "I": "ı"})


def tr_lower(s: str) -> str:
    """Python'un locale-bağımsız str.lower() metodu Türkçe İ harfini
    yanlış çevirir (İ -> 'i' + birleşik nokta işareti, U+0307).
    locale.setlocale() bu davranışı değiştirmez çünkü str.lower()
    locale'e duyarlı değildir. Bu yüzden İ/I harflerini önce elle
    normalize edip sonra lower() çağırıyoruz."""
    return s.translate(_TR_LOWER_MAP).lower()


def to_num(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if s == "" or s == "-":
        return 0
    cleaned = s.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0


def cell_text(value):
    return "" if value is None else str(value).strip()


def find_row(rows, matcher, from_index=0):
    for i in range(from_index, len(rows)):
        cell = rows[i][0] if len(rows[i]) > 0 else None
        text = cell_text(cell)
        if text and matcher(text):
            return i
    return -1


def next_non_empty_row(rows, from_index):
    for i in range(from_index + 1, len(rows)):
        if any(cell_text(c) != "" for c in rows[i]):
            return i
    return -1


def parse_workbook(path: Path) -> dict:
    wb = load_workbook(filename=str(path), data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "saha durum":
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    # Blok 1: Toplam CİK Sayısı / Üretilen Toplam Kabin / Üretilen Toplam MMC
    h1 = find_row(rows, lambda s: "toplam cik sayısı" in tr_lower(s))
    hedef_cik = uretilen_kabin = uretilen_mmc = None
    if h1 >= 0:
        v1 = next_non_empty_row(rows, h1)
        if v1 >= 0:
            r = rows[v1]
            hedef_cik = to_num(r[0] if len(r) > 0 else None)
            uretilen_kabin = to_num(r[1] if len(r) > 1 else None)
            uretilen_mmc = to_num(r[2] if len(r) > 2 else None)
    if hedef_cik is None:
        raise ValueError(
            '"Toplam CİK Sayısı" satırı bulunamadı — dosya formatı beklenenden farklı, veriler okunamadı.'
        )

    # Blok 2: Devreye Alınan Kurum Sayısı / Ay içi Kabin / Ay içi MMC
    h2 = find_row(rows, lambda s: "devreye alınan kurum sayısı" in tr_lower(s))
    devreye_alinan = ay_kabin = ay_mmc = None
    if h2 >= 0:
        v2 = next_non_empty_row(rows, h2)
        if v2 >= 0:
            r = rows[v2]
            devreye_alinan = to_num(r[0] if len(r) > 0 else None)
            ay_kabin = to_num(r[1] if len(r) > 1 else None)
            ay_mmc = to_num(r[2] if len(r) > 2 else None)
    if devreye_alinan is None:
        raise ValueError(
            '"Devreye Alınan Kurum Sayısı" satırı bulunamadı — dosya formatı beklenenden farklı, veriler okunamadı.'
        )

    # Blok 3: Aşama tablosu
    h_table = find_row(rows, lambda s: "aşama" in tr_lower(s) and "durum" in tr_lower(s))
    stages = []
    toplam_row = None
    footnote = ""
    if h_table >= 0:
        for i in range(h_table + 1, len(rows)):
            row = rows[i]
            label = cell_text(row[0] if len(row) > 0 else None)
            if label == "":
                continue
            if tr_lower(label) == "toplam":
                toplam_row = {
                    "cik": to_num(row[1] if len(row) > 1 else None),
                    "kabin": to_num(row[2] if len(row) > 2 else None),
                    "mmc": to_num(row[3] if len(row) > 3 else None),
                }
                continue
            if label.startswith("*"):
                footnote = label
                continue
            stages.append({
                "label": label,
                "cik": to_num(row[1] if len(row) > 1 else None),
                "kabin": to_num(row[2] if len(row) > 2 else None),
                "mmc": to_num(row[3] if len(row) > 3 else None),
                "note": cell_text(row[4] if len(row) > 4 else None),
            })

    return {
        "hedefCik": hedef_cik,
        "uretilenKabin": uretilen_kabin,
        "uretilenMmc": uretilen_mmc,
        "devreyeAlinan": devreye_alinan,
        "ayKabin": ay_kabin,
        "ayMmc": ay_mmc,
        "stages": stages,
        "toplamRow": toplam_row,
        "footnote": footnote,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFile": path.name,
    }


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("saha_durum_operasyon_tablosu.xlsx")
    json_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/dashboard-data.json")

    if not xlsx_path.exists():
        print(f"HATA: {xlsx_path} bulunamadı.", file=sys.stderr)
        sys.exit(1)

    data = parse_workbook(xlsx_path)

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"OK: {json_path} yazıldı ({len(data['stages'])} aşama, hedefCik={data['hedefCik']}).")


if __name__ == "__main__":
    main()
